"""Generate the Unclogme Database Architecture Excel overview."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)

dark_fill = PatternFill('solid', fgColor='1F4E79')
light_blue = PatternFill('solid', fgColor='D6E4F0')
light_green = PatternFill('solid', fgColor='E2EFDA')
light_orange = PatternFill('solid', fgColor='FCE4D6')
light_yellow = PatternFill('solid', fgColor='FFF2CC')
white_fill = PatternFill('solid', fgColor='FFFFFF')
done_fill = PatternFill('solid', fgColor='C6EFCE')
ready_fill = PatternFill('solid', fgColor='FFEB9C')
blocked_fill = PatternFill('solid', fgColor='FFC7CE')
section_fill = PatternFill('solid', fgColor='B4C6E7')

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


# ═══════════════════════════════════════════
# Sheet 1: Architecture Overview
# ═══════════════════════════════════════════
ws = wb.active
ws.title = 'Architecture'

for col, w in [(1,5),(2,25),(3,42),(4,12),(5,30),(6,12),(7,35),(8,30)]:
    ws.column_dimensions[get_column_letter(col)].width = w

# Title
ws.merge_cells('A1:H1')
c = ws['A1']
c.value = 'UNCLOGME - ops.* UNIFIED DATABASE ARCHITECTURE'
c.font = title_font
c.fill = dark_fill
c.alignment = center
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:H2')
c = ws['A2']
c.value = 'Single Source of Truth - All Services Merged Into One Schema | Date: 2026-04-01'
c.font = Font(name='Arial', size=11, italic=True, color='1F4E79')
c.alignment = center

headers = ['#', 'Table Name', 'Description', 'Rows', 'Primary Sources', 'Status', 'Key Columns', 'Notes']

# Build all layers
layers = [
    ('LAYER 1: SOURCE MIRRORS (Raw synced data - READ ONLY - never modify)', light_blue, [
        ['', 'JOBBER (public schema, 11 tables)', '', '', '', '', '', ''],
        ['1', 'jobber_clients', 'Client contact + billing info', '364', 'Jobber API v2', 'SYNCING', 'id, name, email, phone, balance', 'Nightly 6:00 AM UTC'],
        ['2', 'jobber_jobs', 'Work orders (recurring + one-off)', '507', 'Jobber API v2', 'SYNCING', 'id, job_number, client_id, job_type', ''],
        ['3', 'jobber_visits', 'Scheduled service visits', '1,636', 'Jobber API v2', 'SYNCING', 'id, job_id, client_id, visit_status', ''],
        ['4', 'jobber_invoices', 'Billing invoices', '1,583', 'Jobber API v2', 'SYNCING', 'id, invoice_number, total, status', ''],
        ['5', 'jobber_properties', 'Service location addresses', '367', 'Jobber API v2', 'SYNCING', 'id, client_id, street, city', ''],
        ['6', 'jobber_quotes', 'Sales proposals', '171', 'Jobber API v2', 'SYNCING', 'id, quote_number, total', ''],
        ['7', 'jobber_line_items', 'Service line items per job', '0 !!', 'Jobber API v2', 'BROKEN', 'id, job_id, name, unit_price', 'Sync needs repair'],
        ['8', 'jobber_users', 'Employee accounts', '25', 'Jobber API v2', 'SYNCING', 'id, full_name, email, status', ''],
        ['', '', '', '', '', '', '', ''],
        ['', 'AIRTABLE (public schema, 10 tables)', '', '', '', '', '', ''],
        ['9', 'airtable_clients', 'Master client DB (86 columns!)', '185', 'Airtable API', 'SYNCING', 'record_id, client_code, client_name', 'Nightly 7:00 AM UTC'],
        ['10', 'airtable_visits', 'Service visit history', '3,016', 'Airtable API', 'SYNCING', 'record_id, visit_date, service_type, amount', 'Has jobber_visit_id cross-ref'],
        ['11', 'airtable_derm', 'DERM compliance manifests', '868', 'Airtable API', 'SYNCING', 'record_id, white_manifest_num', 'County waste disposal docs'],
        ['12', 'airtable_drivers_team', 'Employee/driver roster', '9', 'Airtable API', 'SYNCING', 'full_name, role, shift, certifications', 'Cross-ref cols exist but empty'],
        ['13', 'airtable_route_creation', 'Route planning', '135', 'Airtable API', 'SYNCING', 'client_record_ids, gt_wanted_date', ''],
        ['14', 'airtable_pre_post_inspection', 'Truck inspections', '239', 'Airtable API', 'SYNCING', 'pre_post, driver, truck, sludge_level', ''],
        ['15', 'airtable_past_due', 'Overdue balances', '45', 'Airtable API', 'SYNCING', 'client_record_ids, amount_due', ''],
        ['', '', '', '', '', '', '', ''],
        ['', 'SAMSARA (samsara schema, 15+ tables)', '', '', '', '', '', ''],
        ['16', 'samsara.vehicles', 'Fleet trucks', '3', 'Samsara API', 'SYNCING', 'id, name, make, model, vin', 'Nightly 6:30 AM UTC'],
        ['17', 'samsara.drivers', 'Fleet drivers', '7', 'Samsara API', 'SYNCING', 'id, name, phone, license_state', ''],
        ['18', 'samsara.addresses', 'Client geofence locations', '192', 'Samsara API', 'SYNCING', 'id, name, lat, lng, formatted_address', '192 verified client sites'],
        ['19', 'samsara.gps_history', 'GPS track points', '535,262', 'Samsara API', 'SYNCING', 'vehicle_id, time, lat, lng', 'Partitioned monthly Oct 25+'],
        ['20', 'samsara.engine_state_events', 'Engine On/Off/Idle', '6,662', 'Samsara API', 'SYNCING', 'vehicle_id, time, value', 'Used for trip reconstruction'],
        ['21', 'samsara.odometer_readings', 'Odometer (meters)', '43,558', 'Samsara API', 'SYNCING', 'vehicle_id, time, value', ''],
        ['22', 'samsara.fuel_readings', 'Fuel level (%)', '18,430', 'Samsara API', 'SYNCING', 'vehicle_id, time, value', ''],
        ['', '', '', '', '', '', '', ''],
        ['', 'FILLOUT (public schema, 3 tables + 4 views)', '', '', '', '', '', ''],
        ['23', 'fillout_pre_shift_inspections', 'Pre-shift truck check form', '94', 'Fillout API', 'SYNCING', 'driver, truck, sludge, gas, photos', 'Nightly 5:00 AM UTC'],
        ['24', 'fillout_post_shift_inspections', 'Post-shift truck check form', '150', 'Fillout API', 'SYNCING', 'driver, truck, sludge, expenses', 'Includes DERM manifest photos'],
        ['', '', '', '', '', '', '', ''],
        ['', 'QUICKBOOKS (public schema, 17 tables)', '', '', '', '', '', ''],
        ['25', 'quickbooks_customers', 'QB customer records', '0', 'QuickBooks API', 'PENDING', 'qb_id, display_name, email, balance', 'Schema deployed, needs API creds'],
        ['26', 'quickbooks_invoices', 'Accounting invoices', '0', 'QuickBooks API', 'PENDING', 'qb_id, doc_number, total_amt', ''],
        ['27', 'quickbooks_payments', 'Payment records', '0', 'QuickBooks API', 'PENDING', 'qb_id, customer_id, total_amt', ''],
        ['28', 'quickbooks_purchases', 'Expenses/bills', '0', 'QuickBooks API', 'PENDING', 'qb_id, total_amt, entity_name', ''],
    ]),
    ('LAYER 2: CANONICAL OPS (The unified database - all queries go here)', light_green, [
        ['1', 'ops.clients', 'THE canonical client table', '491', 'AT + JB + SA merged', 'BUILT', 'id, client_code, canonical_name, status, zone', '118 three-way matches'],
        ['2', 'ops.properties', 'Service locations per client', '~367', 'jobber_properties', 'DESIGNED', 'id, client_id, street, city, postal_code', '1 client = N properties'],
        ['3', 'ops.team_members', 'Employees and drivers', '~25', 'AT team + SA drivers + JB users', 'DESIGNED', 'id, full_name, role, shift, status', 'Cross-refs all 3 systems'],
        ['4', 'ops.vehicles', 'Fleet trucks', '3', 'SA vehicles + AT vehicles', 'DESIGNED', 'id, name, short_code, make, vin, tank_cap', 'Moises, Cloggy, David'],
        ['5', 'ops.jobs', 'Work orders', '~507', 'jobber_jobs', 'DESIGNED', 'id, client_id, job_number, job_type, status', 'Recurring + one-off jobs'],
        ['6', 'ops.visits', 'Service events (CORE TABLE)', '~3,100', 'AT visits + JB visits merged', 'DESIGNED', 'id, client_id, visit_date, service_type', 'MOST IMPORTANT table'],
        ['7', 'ops.invoices', 'Unified invoices', '~1,583', 'JB invoices + QB invoices', 'DESIGNED', 'id, client_id, invoice_number, total', 'Built-in reconciliation'],
        ['8', 'ops.payments', 'Payment records', 'pending', 'QB payments', 'DESIGNED', 'id, client_id, invoice_id, amount', 'Blocked on QB creds'],
        ['9', 'ops.quotes', 'Sales proposals', '~171', 'jobber_quotes', 'DESIGNED', 'id, client_id, quote_number, total', ''],
        ['10', 'ops.line_items', 'Service details per job/inv', 'pending', 'JB line items + QB items', 'DESIGNED', 'id, job_id, name, quantity, unit_price', 'Blocked on JB sync fix'],
        ['11', 'ops.expenses', 'Unified expenses', '~150+', 'QB purchases + Fillout shifts', 'DESIGNED', 'id, expense_date, amount, category', 'Two sources merged'],
        ['12', 'ops.inspections', 'Truck pre/post inspections', '~480', 'Fillout + AT pre_post merged', 'DESIGNED', 'id, vehicle_id, shift_date, type', 'Sludge, fuel, photos'],
        ['13', 'ops.derm_manifests', 'DERM compliance records', '~868', 'airtable_derm', 'DESIGNED', 'id, client_id, manifest_num, sent_to_city', 'County waste disposal'],
        ['14', 'ops.routes', 'Daily route planning', '~135', 'AT route_creation', 'DESIGNED', 'id, client_id, gt_wanted_date, status', ''],
        ['15', 'ops.past_due', 'Outstanding balances', '~45', 'airtable_past_due', 'DESIGNED', 'id, client_id, amount_due, status', ''],
    ]),
    ('LAYER 3: ANALYTICS VIEWS (Real-time dashboards - no data stored, computed live)', light_orange, [
        ['1', 'ops.v_client_health', 'Client compliance + financials', 'live', 'ops.clients + visits + invoices', 'DESIGNED', 'gt_overdue, gdo_expiring, balance', 'One row per active client'],
        ['2', 'ops.v_service_schedule', 'Upcoming services calendar', 'live', 'ops.clients', 'DESIGNED', 'next_visit_date, urgency, zone, price', 'Sorted by urgency'],
        ['3', 'ops.v_compliance_dashboard', 'GDO + GT/CL compliance', 'live', 'ops.clients + derm_manifests', 'DESIGNED', 'gdo_status, gt_compliance, unsent_manifests', 'Urgency-sorted'],
        ['4', 'ops.v_fleet_daily', 'Daily fleet utilization', 'live', 'ops.vehicles + samsara.*', 'DESIGNED', 'miles_driven, engine_hours, fuel', '7-day rolling'],
        ['5', 'ops.v_trips', 'Reconstructed truck trips', 'live', 'samsara engine + gps', 'DESIGNED', 'trip_start/end, duration, destination', 'Engine On/Off based'],
        ['6', 'ops.v_shift_summary', 'Daily shift recap', 'live', 'ops.inspections + team + vehicles', 'DESIGNED', 'driver, sludge_delta, issues, expenses', ''],
        ['7', 'ops.v_revenue_summary', 'Revenue by client/month', 'live', 'ops.invoices + clients', 'DESIGNED', 'billed, outstanding, collected', ''],
    ]),
    ('LAYER 4: REFERENCE (Cross-reference and audit)', light_yellow, [
        ['1', 'ops.entity_map', 'ID cross-reference audit trail', '251', 'Auto-matched', 'BUILT', 'airtable_id, jobber_id, samsara_id, qb_id', 'Matching history'],
    ]),
]

row = 4
for layer_name, layer_fill, data in layers:
    # Layer header
    ws.merge_cells(f'A{row}:H{row}')
    c = ws.cell(row=row, column=1, value=layer_name)
    c.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    c.fill = dark_fill
    c.alignment = center
    ws.row_dimensions[row].height = 28
    row += 1

    # Column headers
    for ci, h in enumerate(headers):
        c = ws.cell(row=row, column=ci+1, value=h)
        c.font = header_font
        c.fill = PatternFill('solid', fgColor='4472C4')
        c.alignment = center
        c.border = thin_border
    row += 1

    # Data
    for dr in data:
        if not dr[1]:
            row += 1
            continue
        is_section = (dr[0] == '' and dr[1])
        for ci, val in enumerate(dr):
            c = ws.cell(row=row, column=ci+1, value=val)
            c.font = bold_font if is_section else normal_font
            c.alignment = left_wrap if ci > 1 else center
            c.border = thin_border
            if is_section:
                c.fill = section_fill
            elif val == 'BUILT':
                c.font = Font(name='Arial', size=10, bold=True, color='006100')
                c.fill = done_fill
            elif val == 'DESIGNED':
                c.font = Font(name='Arial', size=10, bold=True, color='9C5700')
                c.fill = ready_fill
            elif val == 'SYNCING':
                c.font = Font(name='Arial', size=10, color='006100')
                c.fill = done_fill
            elif val == 'PENDING':
                c.font = Font(name='Arial', size=10, color='9C0006')
                c.fill = blocked_fill
            elif val == 'BROKEN':
                c.font = Font(name='Arial', size=10, bold=True, color='9C0006')
                c.fill = blocked_fill
            else:
                c.fill = layer_fill
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1


# ═══════════════════════════════════════════
# Sheet 2: Relationships
# ═══════════════════════════════════════════
ws2 = wb.create_sheet('Relationships')
for col, w in [(1,22),(2,8),(3,22),(4,18),(5,50)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.merge_cells('A1:E1')
c = ws2['A1']
c.value = 'TABLE RELATIONSHIPS (Entity Relationship Diagram)'
c.font = title_font
c.fill = dark_fill
c.alignment = center

rels_headers = ['Parent (1)', 'Type', 'Child (N)', 'FK Column', 'Business Meaning']
rels_data = [
    ['ops.clients', '1:N', 'ops.properties', 'client_id', 'A client has many service locations (e.g., chain restaurants)'],
    ['ops.clients', '1:N', 'ops.jobs', 'client_id', 'A client has many work orders over time'],
    ['ops.clients', '1:N', 'ops.visits', 'client_id', 'A client receives many service visits'],
    ['ops.clients', '1:N', 'ops.invoices', 'client_id', 'A client is billed many invoices'],
    ['ops.clients', '1:N', 'ops.payments', 'client_id', 'A client makes many payments'],
    ['ops.clients', '1:N', 'ops.quotes', 'client_id', 'A client receives many quotes'],
    ['ops.clients', '1:N', 'ops.derm_manifests', 'client_id', 'A client has many DERM compliance records'],
    ['ops.clients', '1:N', 'ops.routes', 'client_id', 'A client appears in many route plans'],
    ['ops.clients', '1:N', 'ops.past_due', 'client_id', 'A client may have past-due entries'],
    ['', '', '', '', ''],
    ['ops.jobs', '1:N', 'ops.visits', 'job_id', 'A recurring job generates many scheduled visits'],
    ['ops.jobs', '1:N', 'ops.invoices', 'job_id', 'A job is invoiced (one or many visit-based invoices)'],
    ['ops.jobs', '1:N', 'ops.line_items', 'job_id', 'A job has detailed service line items'],
    ['', '', '', '', ''],
    ['ops.invoices', '1:N', 'ops.payments', 'invoice_id', 'An invoice can receive multiple partial payments'],
    ['ops.invoices', '1:N', 'ops.line_items', 'invoice_id', 'An invoice lists detailed line items'],
    ['', '', '', '', ''],
    ['ops.properties', '1:N', 'ops.visits', 'property_id', 'Service visits happen at a specific property'],
    ['ops.properties', '1:N', 'ops.jobs', 'property_id', 'Work orders are assigned to a property'],
    ['', '', '', '', ''],
    ['ops.vehicles', '1:N', 'ops.visits', 'vehicle_id', 'A truck performs many visits per shift'],
    ['ops.vehicles', '1:N', 'ops.inspections', 'vehicle_id', 'A truck is inspected pre/post every shift'],
    ['ops.vehicles', '1:N', 'ops.expenses', 'vehicle_id', 'A truck incurs fuel, maintenance, dump expenses'],
    ['', '', '', '', ''],
    ['ops.team_members', '1:N', 'ops.inspections', 'team_member_id', 'A driver submits shift inspections'],
    ['ops.team_members', '1:N', 'ops.expenses', 'team_member_id', 'A driver reports shift expenses'],
]

for ci, h in enumerate(rels_headers):
    c = ws2.cell(row=3, column=ci+1, value=h)
    c.font = header_font
    c.fill = dark_fill
    c.alignment = center
    c.border = thin_border

for ri, rd in enumerate(rels_data):
    for ci, val in enumerate(rd):
        c = ws2.cell(row=ri+4, column=ci+1, value=val)
        c.font = normal_font
        c.alignment = left_wrap if ci == 4 else center
        c.border = thin_border


# ═══════════════════════════════════════════
# Sheet 3: Build Priority
# ═══════════════════════════════════════════
ws3 = wb.create_sheet('Build Priority')
for col, w in [(1,5),(2,22),(3,12),(4,40),(5,12),(6,45)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

ws3.merge_cells('A1:F1')
c = ws3['A1']
c.value = 'BUILD ORDER - Implementation Roadmap'
c.font = title_font
c.fill = dark_fill
c.alignment = center

bh = ['#', 'Table', 'Status', 'Dependencies', 'Effort', 'Notes']
for ci, h in enumerate(bh):
    c = ws3.cell(row=2, column=ci+1, value=h)
    c.font = header_font
    c.fill = dark_fill
    c.alignment = center
    c.border = thin_border

build = [
    ['1', 'ops.clients', 'DONE', 'None (root table)', '~', '491 rows built. 118 three-way matches. LIVE.'],
    ['2', 'ops.vehicles', 'Ready', 'None', '30m', 'Only 3 rows. Simple merge of Samsara + Airtable.'],
    ['3', 'ops.team_members', 'Ready', 'None', '1h', '25 users across 3 systems to deduplicate.'],
    ['4', 'ops.properties', 'Ready', 'ops.clients', '30m', 'Direct from jobber_properties with FK resolution.'],
    ['5', 'ops.jobs', 'Ready', 'ops.clients + properties', '1h', '507 Jobber work orders. Classify service_category.'],
    ['6', 'ops.visits', 'Ready', 'clients + jobs + vehicles', '2-3h', 'CORE TABLE. 3016 AT + 1636 JB = ~3100 merged visits.'],
    ['7', 'ops.invoices', 'Ready', 'ops.clients + jobs', '1h', '1583 Jobber invoices. QB reconciliation when available.'],
    ['8', 'ops.quotes', 'Ready', 'ops.clients + properties', '30m', '171 Jobber quotes. Straightforward.'],
    ['9', 'ops.derm_manifests', 'Ready', 'ops.clients', '1h', '868 DERM records. Resolve client cross-refs from Airtable.'],
    ['10', 'ops.inspections', 'Ready', 'vehicles + team_members', '2h', 'Merge 94+150 Fillout + 239 Airtable inspections.'],
    ['11', 'ops.expenses', 'Partial', 'vehicles + team_members', '1h', 'Fillout expenses now. QB purchases later.'],
    ['12', 'ops.routes', 'Ready', 'ops.clients', '30m', '135 Airtable route plans. Simple.'],
    ['13', 'ops.past_due', 'Ready', 'ops.clients', '30m', '45 overdue accounts. Simple.'],
    ['14', 'ops.line_items', 'Blocked', 'ops.jobs + invoices', '--', 'Jobber sync returning 0 rows. Fix sync first.'],
    ['15', 'ops.payments', 'Blocked', 'ops.invoices', '--', 'Needs QuickBooks API credentials from Fred.'],
    ['', '', '', '', '', ''],
    ['', 'VIEWS', '', '', '', ''],
    ['16', 'v_service_schedule', 'Ready', 'ops.clients only', '15m', 'Can build immediately.'],
    ['17', 'v_compliance_dashboard', 'Ready', 'clients + derm', '30m', 'After derm_manifests built.'],
    ['18', 'v_client_health', 'Ready', 'clients + visits + invoices', '30m', 'After visits + invoices built.'],
    ['19', 'v_fleet_daily', 'Ready', 'vehicles + samsara.*', '30m', 'Uses Samsara telemetry directly.'],
    ['20', 'v_trips', 'Ready', 'vehicles + samsara.*', '30m', 'Engine On/Off trip reconstruction.'],
    ['21', 'v_shift_summary', 'Ready', 'inspections + team + vehicles', '30m', 'After inspections built.'],
    ['22', 'v_revenue_summary', 'Ready', 'invoices + clients', '15m', 'After invoices built.'],
]

for ri, rd in enumerate(build):
    for ci, val in enumerate(rd):
        c = ws3.cell(row=ri+3, column=ci+1, value=val)
        c.font = bold_font if rd[0] == '' else normal_font
        c.alignment = left_wrap if ci > 2 else center
        c.border = thin_border
        if rd[0] == '' and rd[1]:
            c.fill = section_fill
            c.font = bold_font
        elif val == 'DONE':
            c.font = Font(name='Arial', size=10, bold=True, color='006100')
            c.fill = done_fill
        elif val == 'Ready':
            c.font = Font(name='Arial', size=10, color='9C5700')
            c.fill = ready_fill
        elif val in ('Blocked', 'Partial'):
            c.font = Font(name='Arial', size=10, color='9C0006')
            c.fill = blocked_fill

out = r'C:\Users\FRED\Desktop\Virtrify\Yannick\Claude\Slack\DERM\Unclogme_Database_Architecture.xlsx'
wb.save(out)
print(f'Saved to {out}')
