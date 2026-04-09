"""Generate the Unclogme Database Architecture Excel overview v2.0
   Updated: No QuickBooks, Ramp for expenses, 4 trucks (Goliath!), new Supabase project."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)

dark_fill = PatternFill('solid', fgColor='1F4E79')
light_blue = PatternFill('solid', fgColor='D6E4F0')
light_green = PatternFill('solid', fgColor='E2EFDA')
light_orange = PatternFill('solid', fgColor='FCE4D6')
light_yellow = PatternFill('solid', fgColor='FFF2CC')
white_fill = PatternFill('solid', fgColor='FFFFFF')
done_fill = PatternFill('solid', fgColor='C6EFCE')
ready_fill = PatternFill('solid', fgColor='FFEB9C')
blocked_fill = PatternFill('solid', fgColor='FFC7CE')
section_fill = PatternFill('solid', fgColor='B4C6E7')

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_headers(ws, row, headers):
    for ci, h in enumerate(headers):
        c = ws.cell(row=row, column=ci+1, value=h)
        c.font = header_font
        c.fill = dark_fill
        c.alignment = center
        c.border = thin_border


# ═══════════════════════════════════════════
# Sheet 1: Architecture Overview
# ═══════════════════════════════════════════
ws = wb.active
ws.title = 'Architecture'

for col, w in [(1,5),(2,25),(3,42),(4,12),(5,30),(6,12),(7,35),(8,35)]:
    ws.column_dimensions[get_column_letter(col)].width = w

# Title
ws.merge_cells('A1:H1')
c = ws['A1']
c.value = 'UNCLOGME - ops.* UNIFIED DATABASE ARCHITECTURE v2.0'
c.font = title_font; c.fill = dark_fill; c.alignment = center
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:H2')
c = ws['A2']
c.value = 'Single Source of Truth | Deploy to NEW Supabase Project | No QuickBooks | Date: 2026-04-01'
c.font = Font(name='Arial', size=11, italic=True, color='1F4E79')
c.alignment = center

headers = ['#', 'Table Name', 'Description', 'Rows', 'Primary Sources', 'Status', 'Key Columns', 'Notes']

# All layers data
layers = [
    ('LAYER 1: SOURCE MIRRORS (Existing Supabase - raw synced data - READ ONLY)', light_blue, [
        ['', 'JOBBER (public schema, 11 tables)', '', '', '', '', '', ''],
        ['1', 'jobber_clients', 'Client contact + billing info', '364', 'Jobber API', 'SYNCING', 'id, name, email, phone, balance', 'Nightly 6:00 AM UTC'],
        ['2', 'jobber_jobs', 'Work orders (recurring + one-off)', '507', 'Jobber API', 'SYNCING', 'id, job_number, client_id, job_type', ''],
        ['3', 'jobber_visits', 'Scheduled service visits', '1,636', 'Jobber API', 'SYNCING', 'id, job_id, client_id, visit_status', ''],
        ['4', 'jobber_invoices', 'Invoices + PAYMENT tracking', '1,583', 'Jobber API', 'SYNCING', 'id, invoice_number, total, paid_at', 'Jobber = payment authority'],
        ['5', 'jobber_properties', 'Service location addresses', '367', 'Jobber API', 'SYNCING', 'id, client_id, street, city', ''],
        ['6', 'jobber_quotes', 'Sales proposals', '171', 'Jobber API', 'SYNCING', 'id, quote_number, total', '5 high-value unsigned'],
        ['7', 'jobber_line_items', 'Service line items per job', '0 !!', 'Jobber API', 'BROKEN', 'id, job_id, name, unit_price', 'SYNC BROKEN - fix needed'],
        ['8', 'jobber_users', 'Employee accounts', '25', 'Jobber API', 'SYNCING', 'id, full_name, email, status', ''],
        ['', '', '', '', '', '', '', ''],
        ['', 'AIRTABLE (public schema, 10 tables)', '', '', '', '', '', ''],
        ['9', 'airtable_clients', 'CRM master (86 columns!)', '185', 'Airtable API', 'SYNCING', 'record_id, client_code, client_name', 'Nightly 7:00 AM UTC'],
        ['10', 'airtable_visits', 'Service visit history', '3,016', 'Airtable API', 'SYNCING', 'record_id, visit_date, service_type', 'Has jobber_visit_id xref'],
        ['11', 'airtable_derm', 'DERM compliance manifests', '868', 'Airtable API', 'SYNCING', 'record_id, white_manifest_num', 'County-required docs'],
        ['12', 'airtable_drivers_team', 'Employee/driver roster', '9', 'Airtable API', 'SYNCING', 'full_name, role, shift, certifications', ''],
        ['13', 'airtable_route_creation', 'Route planning', '135', 'Airtable API', 'SYNCING', 'client_record_ids, gt_wanted_date', ''],
        ['14', 'airtable_pre_post_inspection', 'Truck inspections (legacy)', '239', 'Airtable API', 'SYNCING', 'pre_post, driver, truck, sludge_level', 'Migrating to Fillout'],
        ['15', 'airtable_past_due', 'Overdue balances', '45', 'Airtable API', 'SYNCING', 'client_record_ids, amount_due', ''],
        ['', '', '', '', '', '', '', ''],
        ['', 'SAMSARA (samsara schema, 15+ tables)', '', '', '', '', '', ''],
        ['16', 'samsara.vehicles', 'Fleet trucks (3 of 4)', '3', 'Samsara API', 'SYNCING', 'id, name, make, model, vin', 'Goliath not on Samsara'],
        ['17', 'samsara.drivers', 'Fleet drivers', '7', 'Samsara API', 'SYNCING', 'id, name, phone, license_state', 'Nightly 6:30 AM UTC'],
        ['18', 'samsara.addresses', 'Client geofence locations', '192', 'Samsara API', 'SYNCING', 'id, name, lat, lng, formatted_address', '192 verified client sites'],
        ['19', 'samsara.gps_history', 'GPS track points', '535K+', 'Samsara API', 'SYNCING', 'vehicle_id, time, lat, lng', 'Partitioned monthly'],
        ['20', 'samsara.engine_state_events', 'Engine On/Off/Idle', '6,662', 'Samsara API', 'SYNCING', 'vehicle_id, time, value', 'Trip reconstruction'],
        ['', '', '', '', '', '', '', ''],
        ['', 'FILLOUT (public schema, 3 tables + 4 views)', '', '', '', '', '', ''],
        ['21', 'fillout_pre_shift_inspections', 'Pre-shift truck check', '94', 'Fillout API', 'SYNCING', 'driver, truck, sludge, gas, photos', 'Nightly 5:00 AM UTC'],
        ['22', 'fillout_post_shift_inspections', 'Post-shift truck check', '150', 'Fillout API', 'SYNCING', 'driver, truck, sludge, expenses', 'Includes DERM photos'],
        ['', '', '', '', '', '', '', ''],
        ['', 'RAMP (expense management - integration pending)', '', '', '', '', '', ''],
        ['23', 'ramp_transactions', 'Corporate card transactions', 'TBD', 'Ramp API', 'PLANNED', 'transaction_id, amount, merchant', 'Needs Ramp API access'],
    ]),
    ('LAYER 2: CANONICAL OPS (New Supabase - the unified database)', light_green, [
        ['1', 'ops.clients', 'THE canonical client table', '~491', 'AT + JB + SA merged', 'DESIGNED', 'id, client_code, canonical_name, status, zone', '118 three-way matches'],
        ['2', 'ops.properties', 'Service locations per client', '~367', 'jobber_properties', 'DESIGNED', 'id, client_id, street, city, postal_code', 'La Granja 5+, Carrot Express 4+'],
        ['3', 'ops.team_members', 'Employees and drivers', '~10', 'AT team + SA drivers + JB users', 'DESIGNED', 'id, full_name, role, shift, access_level', 'Dev > Office > Field access'],
        ['4', 'ops.vehicles', 'Fleet trucks (ALL 4)', '4', 'SA vehicles + AT vehicles', 'DESIGNED', 'id, name, short_code, tank_capacity', 'Cloggy, David, Goliath, Moise'],
        ['5', 'ops.jobs', 'Work orders', '~507', 'jobber_jobs', 'DESIGNED', 'id, client_id, job_number, job_type', 'Recurring + one-off'],
        ['6', 'ops.visits', 'Service events (CORE TABLE)', '~3,100', 'AT visits + JB visits merged', 'DESIGNED', 'id, client_id, visit_date, service_type', 'THE most important table'],
        ['7', 'ops.invoices', 'Invoices + payment status', '~1,583', 'jobber_invoices', 'DESIGNED', 'id, client_id, invoice_number, paid_at', 'A/R: ~$132K outstanding'],
        ['8', 'ops.quotes', 'Sales proposals', '~171', 'jobber_quotes', 'DESIGNED', 'id, client_id, quote_number, total', '5 high-value unsigned'],
        ['9', 'ops.line_items', 'Service details per job', 'pending', 'jobber_line_items', 'DESIGNED', 'id, job_id, name, quantity, unit_price', 'Blocked on JB sync fix'],
        ['10', 'ops.expenses', 'Unified expenses', '~150+', 'Ramp + Fillout shifts', 'DESIGNED', 'id, expense_date, amount, category', 'Corporate cards + driver reports'],
        ['11', 'ops.inspections', 'Truck pre/post inspections', '~480', 'Fillout + AT merged', 'DESIGNED', 'id, vehicle_id, shift_date, sludge', 'Sludge delta = waste collected'],
        ['12', 'ops.derm_manifests', 'DERM compliance records', '~868', 'airtable_derm', 'DESIGNED', 'id, client_id, manifest_num, sent_to_city', 'Fines: $500-$3,000'],
        ['13', 'ops.routes', 'Route planning', '~135', 'AT route_creation', 'DESIGNED', 'id, client_id, gt_wanted_date, status', ''],
        ['14', 'ops.past_due', 'Outstanding balances', '~45', 'airtable_past_due', 'DESIGNED', 'id, client_id, amount_due, status', '77 late clients = $130K ARR'],
        ['15', 'ops.entity_map', 'Cross-reference audit', '~251', 'Auto-matched', 'DESIGNED', 'airtable_id, jobber_id, samsara_id', 'Matching history'],
    ]),
    ('LAYER 3: ANALYTICS VIEWS (Computed live - no data stored)', light_orange, [
        ['1', 'ops.v_client_health', 'Client compliance + financials', 'live', 'clients + visits + invoices', 'DESIGNED', 'gt_overdue, gdo_expiring, balance', 'One row per active client'],
        ['2', 'ops.v_service_schedule', 'Upcoming services calendar', 'live', 'ops.clients', 'DESIGNED', 'next_visit_date, urgency, zone, price', 'CRITICAL > OVERDUE > DUE_SOON'],
        ['3', 'ops.v_compliance_dashboard', 'GDO + DERM compliance', 'live', 'clients + derm_manifests', 'DESIGNED', 'gdo_status, gt_compliance, exceeds_90d', 'DERM max 90 days. ~30% at risk'],
        ['4', 'ops.v_fleet_daily', 'Daily fleet utilization', 'live', 'vehicles + samsara telemetry', 'DESIGNED', 'miles_driven, engine_hours, fuel', 'Placeholder until source DB linked'],
        ['5', 'ops.v_trips', 'Reconstructed truck trips', 'live', 'samsara engine + gps', 'DESIGNED', 'trip_start/end, duration, destination', 'Engine On/Off based'],
        ['6', 'ops.v_shift_summary', 'Daily shift recap', 'live', 'inspections + team + vehicles', 'DESIGNED', 'sludge_collected, issues, expenses', 'Sludge delta = POST - PRE'],
        ['7', 'ops.v_revenue_summary', 'Revenue by client/month', 'live', 'invoices + clients', 'DESIGNED', 'billed, collected, outstanding', 'Current: $674K/yr, target: $200K/mo'],
    ]),
]

row = 4
for layer_name, layer_fill, data in layers:
    ws.merge_cells(f'A{row}:H{row}')
    c = ws.cell(row=row, column=1, value=layer_name)
    c.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    c.fill = dark_fill; c.alignment = center
    ws.row_dimensions[row].height = 28
    row += 1
    write_headers(ws, row, headers)
    row += 1

    for dr in data:
        if not dr[1]:
            row += 1
            continue
        is_section = (dr[0] == '' and dr[1])
        for ci, val in enumerate(dr):
            c = ws.cell(row=row, column=ci+1, value=val)
            c.font = bold_font if is_section else normal_font
            c.alignment = left_wrap if ci > 1 else center
            c.border = thin_border
            if is_section: c.fill = section_fill
            elif val == 'BUILT': c.font = Font(name='Arial', size=10, bold=True, color='006100'); c.fill = done_fill
            elif val == 'DESIGNED': c.font = Font(name='Arial', size=10, bold=True, color='9C5700'); c.fill = ready_fill
            elif val == 'SYNCING': c.font = Font(name='Arial', size=10, color='006100'); c.fill = done_fill
            elif val == 'PLANNED': c.font = Font(name='Arial', size=10, color='1F4E79'); c.fill = PatternFill('solid', fgColor='D6E4F0')
            elif val in ('PENDING', 'BROKEN'): c.font = Font(name='Arial', size=10, bold=True, color='9C0006'); c.fill = blocked_fill
            else: c.fill = layer_fill
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1


# ═══════════════════════════════════════════
# Sheet 2: Relationships
# ═══════════════════════════════════════════
ws2 = wb.create_sheet('Relationships')
for col, w in [(1,22),(2,8),(3,22),(4,18),(5,55)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.merge_cells('A1:E1')
c = ws2['A1']
c.value = 'TABLE RELATIONSHIPS (Entity Relationship Diagram)'
c.font = title_font; c.fill = dark_fill; c.alignment = center

write_headers(ws2, 3, ['Parent (1)', 'Type', 'Child (N)', 'FK Column', 'Business Meaning'])

rels = [
    ['ops.clients', '1:N', 'ops.properties', 'client_id', 'A client has many service locations (chain restaurants)'],
    ['ops.clients', '1:N', 'ops.jobs', 'client_id', 'A client has many work orders over time'],
    ['ops.clients', '1:N', 'ops.visits', 'client_id', 'A client receives many service visits'],
    ['ops.clients', '1:N', 'ops.invoices', 'client_id', 'A client is billed many invoices'],
    ['ops.clients', '1:N', 'ops.quotes', 'client_id', 'A client receives many quotes'],
    ['ops.clients', '1:N', 'ops.derm_manifests', 'client_id', 'A client has many DERM compliance records'],
    ['ops.clients', '1:N', 'ops.routes', 'client_id', 'A client appears in route planning'],
    ['ops.clients', '1:N', 'ops.past_due', 'client_id', 'A client may have past-due entries'],
    ['', '', '', '', ''],
    ['ops.jobs', '1:N', 'ops.visits', 'job_id', 'A recurring job generates many scheduled visits'],
    ['ops.jobs', '1:N', 'ops.invoices', 'job_id', 'A job is invoiced (per-visit or fixed-price)'],
    ['ops.jobs', '1:N', 'ops.line_items', 'job_id', 'A job has detailed service line items'],
    ['', '', '', '', ''],
    ['ops.invoices', '1:N', 'ops.line_items', 'invoice_id', 'An invoice details service items'],
    ['', '', '', '', ''],
    ['ops.properties', '1:N', 'ops.visits', 'property_id', 'Visits happen at a specific property'],
    ['ops.properties', '1:N', 'ops.jobs', 'property_id', 'Work orders are for a specific property'],
    ['', '', '', '', ''],
    ['ops.vehicles', '1:N', 'ops.visits', 'vehicle_id', 'A truck performs many visits per shift'],
    ['ops.vehicles', '1:N', 'ops.inspections', 'vehicle_id', 'A truck is inspected each shift'],
    ['ops.vehicles', '1:N', 'ops.expenses', 'vehicle_id', 'A truck incurs fuel/maintenance expenses'],
    ['', '', '', '', ''],
    ['ops.team_members', '1:N', 'ops.inspections', 'team_member_id', 'A driver submits shift inspections'],
    ['ops.team_members', '1:N', 'ops.expenses', 'team_member_id', 'A driver reports shift expenses'],
]

for ri, rd in enumerate(rels):
    for ci, val in enumerate(rd):
        c = ws2.cell(row=ri+4, column=ci+1, value=val)
        c.font = normal_font; c.alignment = left_wrap if ci == 4 else center; c.border = thin_border


# ═══════════════════════════════════════════
# Sheet 3: Build Priority
# ═══════════════════════════════════════════
ws3 = wb.create_sheet('Build Priority')
for col, w in [(1,5),(2,22),(3,12),(4,40),(5,12),(6,50)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

ws3.merge_cells('A1:F1')
c = ws3['A1']
c.value = 'BUILD ORDER - Implementation Roadmap'
c.font = title_font; c.fill = dark_fill; c.alignment = center

write_headers(ws3, 2, ['#', 'Table', 'Status', 'Dependencies', 'Effort', 'Notes'])

build = [
    ['', 'PHASE 1: Foundation (no dependencies)', '', '', '', ''],
    ['1', 'ops.clients', 'Ready', 'None (root table)', '2h', '491 rows. Merge AT+JB+SA. Already prototyped on existing Supabase.'],
    ['2', 'ops.vehicles', 'Ready', 'None', '30m', '4 rows: Cloggy, David, Goliath, Moise. Goliath added manually.'],
    ['3', 'ops.team_members', 'Ready', 'None', '1h', '~10 active. Deduplicate AT team (9) + SA drivers (7) + JB users (25).'],
    ['4', 'ops.entity_map', 'Ready', 'None', '1h', 'Audit trail. 251 auto-matched cross-references.'],
    ['', '', '', '', '', ''],
    ['', 'PHASE 2: Core Operations (depend on Phase 1)', '', '', '', ''],
    ['5', 'ops.properties', 'Ready', 'ops.clients', '30m', '367 Jobber properties. Resolve client_id via entity_map.'],
    ['6', 'ops.jobs', 'Ready', 'ops.clients + properties', '1h', '507 work orders. Classify service_category from title.'],
    ['7', 'ops.visits', 'Ready', 'clients + jobs + vehicles', '3h', 'CORE TABLE. Merge 3016 AT + 1636 JB = ~3100 visits. Complex matching.'],
    ['8', 'ops.inspections', 'Ready', 'vehicles + team_members', '2h', 'Merge 94+150 Fillout + 239 AT inspections. Dedup overlaps.'],
    ['', '', '', '', '', ''],
    ['', 'PHASE 3: Financial + Compliance (depend on Phase 2)', '', '', '', ''],
    ['9', 'ops.invoices', 'Ready', 'ops.clients + jobs', '1h', '1583 Jobber invoices. Payment tracked via paid_at.'],
    ['10', 'ops.quotes', 'Ready', 'ops.clients + properties', '30m', '171 quotes. 5 high-value unsigned: Casa Neos, Chima, etc.'],
    ['11', 'ops.derm_manifests', 'Ready', 'ops.clients', '1h', '868 records. Resolve client via AT record_id cross-ref.'],
    ['12', 'ops.routes', 'Ready', 'ops.clients', '30m', '135 route plans from Airtable.'],
    ['13', 'ops.past_due', 'Ready', 'ops.clients', '30m', '45 overdue accounts. 77 clients late = $130K locked ARR.'],
    ['', '', '', '', '', ''],
    ['', 'PHASE 4: Extended (depend on Phase 3 + external fixes)', '', '', '', ''],
    ['14', 'ops.line_items', 'Blocked', 'ops.jobs + invoices', '--', 'Jobber sync returning 0 rows. Fix Jobber API sync first.'],
    ['15', 'ops.expenses', 'Partial', 'vehicles + team_members', '1h', 'Fillout expenses now. Ramp API integration later.'],
    ['', '', '', '', '', ''],
    ['', 'VIEWS (build after their dependencies)', '', '', '', ''],
    ['16', 'v_service_schedule', 'Ready', 'ops.clients only', '15m', 'Can build immediately after Phase 1.'],
    ['17', 'v_compliance_dashboard', 'Ready', 'clients + derm', '30m', 'After DERM manifests. DERM 90-day max tracking.'],
    ['18', 'v_client_health', 'Ready', 'clients + visits + invoices', '30m', 'After Phase 3.'],
    ['19', 'v_shift_summary', 'Ready', 'inspections + team + vehicles', '30m', 'After Phase 2. Sludge delta tracking.'],
    ['20', 'v_revenue_summary', 'Ready', 'invoices + clients', '15m', 'After invoices built.'],
    ['21', 'v_fleet_daily', 'Ready', 'vehicles + source DB link', '30m', 'Needs connection to existing Supabase for Samsara data.'],
    ['22', 'v_trips', 'Ready', 'vehicles + source DB link', '30m', 'Engine On/Off trip reconstruction.'],
]

for ri, rd in enumerate(build):
    for ci, val in enumerate(rd):
        c = ws3.cell(row=ri+3, column=ci+1, value=val)
        c.font = bold_font if rd[0] == '' else normal_font
        c.alignment = left_wrap if ci > 2 else center
        c.border = thin_border
        if rd[0] == '' and rd[1]:
            c.fill = section_fill; c.font = bold_font
        elif val == 'Ready':
            c.font = Font(name='Arial', size=10, color='006100'); c.fill = done_fill
        elif val in ('Blocked', 'Partial'):
            c.font = Font(name='Arial', size=10, color='9C0006'); c.fill = blocked_fill


# ═══════════════════════════════════════════
# Sheet 4: Fleet & Team Reference
# ═══════════════════════════════════════════
ws4 = wb.create_sheet('Fleet & Team')
for col, w in [(1,12),(2,22),(3,8),(4,12),(5,15),(6,25),(7,15),(8,20)]:
    ws4.column_dimensions[get_column_letter(col)].width = w

ws4.merge_cells('A1:H1')
c = ws4['A1']
c.value = 'FLEET & TEAM REFERENCE'
c.font = title_font; c.fill = dark_fill; c.alignment = center

# Fleet
ws4.merge_cells('A3:H3')
ws4['A3'] = 'THE FLEET (4 trucks - currently at 17% capacity - can serve 700+ clients)'
ws4['A3'].font = Font(name='Arial', size=12, bold=True, color='1F4E79')

write_headers(ws4, 4, ['Name', 'Vehicle', 'Year', 'Capacity', 'Primary Use', 'Short Code', 'On Samsara', 'Notes'])
fleet = [
    ['Cloggy', 'Toyota Tundra', '2020', '126 gal', 'Day jobs, small residential', 'TOY', 'Yes', ''],
    ['David', 'International MA025', '2017', '1,800 gal', 'Night commercial', 'INT', 'Yes', ''],
    ['Goliath', 'Peterbilt 579', '2019', '4,800 gal', 'Large commercial', 'PET', 'No', 'NOT on Samsara GPS'],
    ['Moise', 'Kenworth T880', '2023', '9,000 gal', 'Large commercial (newest)', 'KEN', 'Yes', '$360K custom build, Jan 2026'],
]
for ri, rd in enumerate(fleet):
    for ci, val in enumerate(rd):
        c = ws4.cell(row=ri+5, column=ci+1, value=val)
        c.font = normal_font; c.border = thin_border; c.alignment = center
        if val == 'No':
            c.font = Font(name='Arial', size=10, bold=True, color='9C0006'); c.fill = blocked_fill

# Team
ws4.merge_cells('A11:H11')
ws4['A11'] = 'THE TEAM (10 members - 3 access levels)'
ws4['A11'].font = Font(name='Arial', size=12, bold=True, color='1F4E79')

write_headers(ws4, 12, ['Name', 'Role', '', 'Shift', 'Access Level', 'Contact', '', 'Notes'])
team = [
    ['Yannick Ayache', 'Founder / Owner', '', '', 'Dev', 'yannick@ayache.com', '', 'Strategy & growth'],
    ['Fred Zerpa', 'Admin & Tech Director', '', '', 'Dev', 'fredzerpa@gmail.com', '', 'Builds everything tech'],
    ['Aaron Azoulay', 'Operations Manager', '', '', 'Office', 'aaron@unclogme.com', '', 'Scheduling, dispatch, clients'],
    ['Diego Hernandez', 'Office Manager', '', '', 'Office', 'contact@unclogme.com', '', 'Schedules, invoicing, comms'],
    ['Andres Machado', 'Team Lead & Fleet Mgr', '', 'Both', 'Field', '', '', 'Hired Aug 2025'],
    ['Pablo', 'Technician', '', 'Night', 'Field', '', '', ''],
    ['Brian', 'Technician', '', 'Both', 'Field', '', '', ''],
    ['Kevis Bell', 'Technician', '', 'Both', 'Field', '', '', ''],
    ['Ishad', 'Technician', '', 'Both', 'Field', '', '', ''],
    ['Grecia', 'Part-Time Tech', '', 'Night', 'Field', '', '', ''],
]
for ri, rd in enumerate(team):
    for ci, val in enumerate(rd):
        c = ws4.cell(row=ri+13, column=ci+1, value=val)
        c.font = normal_font; c.border = thin_border; c.alignment = center
        if val == 'Dev': c.fill = PatternFill('solid', fgColor='D6E4F0')
        elif val == 'Office': c.fill = ready_fill
        elif val == 'Field': c.fill = PatternFill('solid', fgColor='E2EFDA')


out = r'C:\Users\FRED\Desktop\Virtrify\Yannick\Claude\Slack\DERM\Unclogme_Database_Architecture.xlsx'
wb.save(out)
print(f'Saved: {out}')
