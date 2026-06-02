import json
import os
import sys
from pathlib import Path

# Force UTF-8 stdout to avoid cp1252 errors on Windows
os.environ['PYTHONIOENCODING'] = 'utf-8'
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[2]
SRC = BASE / 'docs' / 'audits' / '2026-05-19' / 'recurring_jobs.json'
OUT = BASE / 'docs' / 'audits' / '2026-05-19' / 'Recurring_Jobs_2026-05-19.xlsx'

with open(SRC, 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = 'Recurring Jobs'

# Styling
HEADER_FILL = PatternFill('solid', start_color='E7E5E4')
HEADER_FONT = Font(name='Arial', bold=True, color='1F2937', size=11)
CLIENT_FILL = PatternFill('solid', start_color='FEE9DF')
CLIENT_FONT = Font(name='Arial', bold=True, color='9A3412', size=11)
SUBTOTAL_FILL = PatternFill('solid', start_color='F5F5F4')
SUBTOTAL_FONT = Font(name='Arial', bold=True, color='44403C', italic=True, size=10)
ROW_FONT = Font(name='Arial', size=10, color='1F2937')
MUTED_FONT = Font(name='Arial', size=10, color='6B7280')
THIN = Side(style='thin', color='E7E5E4')
BORDER = Border(bottom=THIN)

CURRENCY = '"$"#,##0.00'
QTY = '#,##0.##'

HEADERS = ['Client Code', 'Client Name', 'Job #', 'Job Title', 'Status', 'Start', 'Service / Product', 'Description', 'Qty', 'Unit Price', 'Line Total']

# Header row
ws.append(HEADERS)
for col_i, _ in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_i)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = BORDER

ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

# Sort groups: client_code asc, None at end
groups = data['groups']
def sort_key(g):
    code = g.get('client_code')
    return (1 if code is None else 0, code or '', g.get('client_name', '') or '')
groups.sort(key=sort_key)

row = 2
for g in groups:
    code = g.get('client_code') or ''
    name = g.get('client_name') or '(no name)'
    client_start_row = row
    has_any_line_item = False

    # Sort jobs by jobNumber asc
    g['jobs'].sort(key=lambda j: (j.get('jobNumber') or 0))

    for j in g['jobs']:
        job_number = j.get('jobNumber') or ''
        title = j.get('title') or ''
        status = j.get('jobStatus') or ''
        start_at = (j.get('startAt') or '')[:10]
        line_items = (j.get('lineItems') or {}).get('nodes') or []

        if not line_items:
            ws.cell(row=row, column=1, value=code).font = ROW_FONT
            ws.cell(row=row, column=2, value=name).font = ROW_FONT
            ws.cell(row=row, column=3, value=job_number).font = ROW_FONT
            ws.cell(row=row, column=4, value=title).font = ROW_FONT
            ws.cell(row=row, column=5, value=status).font = MUTED_FONT
            ws.cell(row=row, column=6, value=start_at).font = MUTED_FONT
            ws.cell(row=row, column=7, value='(no line items)').font = MUTED_FONT
            for c in range(1, 12):
                ws.cell(row=row, column=c).border = BORDER
            row += 1
            continue

        for li in line_items:
            ws.cell(row=row, column=1, value=code).font = ROW_FONT
            ws.cell(row=row, column=2, value=name).font = ROW_FONT
            ws.cell(row=row, column=3, value=job_number).font = ROW_FONT
            ws.cell(row=row, column=4, value=title).font = ROW_FONT
            ws.cell(row=row, column=5, value=status).font = MUTED_FONT
            ws.cell(row=row, column=6, value=start_at).font = MUTED_FONT
            ws.cell(row=row, column=7, value=li.get('name') or '').font = ROW_FONT
            ws.cell(row=row, column=8, value=li.get('description') or '').font = MUTED_FONT
            qty_cell = ws.cell(row=row, column=9, value=li.get('quantity'))
            qty_cell.font = ROW_FONT
            qty_cell.number_format = QTY
            up_cell = ws.cell(row=row, column=10, value=li.get('unitPrice'))
            up_cell.font = ROW_FONT
            up_cell.number_format = CURRENCY
            total_cell = ws.cell(row=row, column=11, value=li.get('totalPrice'))
            total_cell.font = ROW_FONT
            total_cell.number_format = CURRENCY
            for c in range(1, 12):
                ws.cell(row=row, column=c).border = BORDER
            row += 1
            has_any_line_item = True

    # Client subtotal row
    if row > client_start_row:
        subtotal_row = row
        ws.cell(row=subtotal_row, column=2, value=f'{name} — total').font = SUBTOTAL_FONT
        if has_any_line_item:
            formula = f'=SUM(K{client_start_row}:K{subtotal_row - 1})'
            sub_cell = ws.cell(row=subtotal_row, column=11, value=formula)
            sub_cell.font = SUBTOTAL_FONT
            sub_cell.number_format = CURRENCY
        for c in range(1, 12):
            ws.cell(row=subtotal_row, column=c).fill = SUBTOTAL_FILL
            ws.cell(row=subtotal_row, column=c).border = BORDER
        row += 1
        # blank spacer row
        row += 1

# Column widths
widths = {1: 12, 2: 32, 3: 12, 4: 36, 5: 18, 6: 12, 7: 28, 8: 38, 9: 8, 10: 12, 11: 14}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Summary sheet
ws2 = wb.create_sheet('Summary')
ws2.append(['Metric', 'Value'])
for c in range(1, 3):
    cell = ws2.cell(row=1, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
ws2.append(['Total recurring jobs', data['total_jobs']])
ws2.append(['Total clients', data['client_count']])
total_lines = sum(len((j.get('lineItems') or {}).get('nodes') or []) for g in groups for j in g['jobs'])
ws2.append(['Total line items', total_lines])
ws2.append(['Generated', data.get('generated_at', '')])
ws2.column_dimensions['A'].width = 26
ws2.column_dimensions['B'].width = 30

wb.save(OUT)
print(f'Written: {OUT}')
print(f'Rows: {row - 1} on Recurring Jobs sheet')
